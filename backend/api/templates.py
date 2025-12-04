"""
模板相关 API 路由
处理表单模板的创建、查看、编辑等操作
"""
from fastapi import APIRouter, Depends, HTTPException, status, UploadFile, File
from sqlalchemy.orm import Session
from pydantic import BaseModel, Field
from typing import List, Optional
from datetime import datetime
import openpyxl
from openpyxl.cell import Cell
from openpyxl.worksheet.datavalidation import DataValidation
import io
from backend.utils import ensure_utc

from backend.database.db_config import get_db_session
from backend.database.models import TemplateForm, TemplateFormField, Secretary
from backend.api.auth import get_current_user
from backend.logger import get_logger
from backend.utils.template_utils import create_template_core, update_template_core, ALLOWED_TYPES

logger = get_logger(__name__)

router = APIRouter()


# ==================== Pydantic 模型 ====================

class FieldRequest(BaseModel):
    """字段请求"""
    display_name: str = Field(..., min_length=1, max_length=100)
    # validation_rule JSON: see FieldResponse.validation_rule for structure
    validation_rule: Optional[dict] = None
    ord: int = Field(..., ge=0, description="字段顺序")
    class Config:
        extra = 'forbid'


class FieldResponse(BaseModel):
    """字段响应"""
    id: int
    display_name: str
    validation_rule: Optional[dict] = None
    ord: int


class TemplateResponse(BaseModel):
    """模板响应"""
    id: int
    name: str
    description: Optional[str]
    created_at: datetime
    field_count: int = 0


class TemplateDetailResponse(BaseModel):
    """模板详情响应"""
    id: int
    name: str
    description: Optional[str]
    created_at: datetime
    fields: List[FieldResponse]


class CreateTemplateRequest(BaseModel):
    """创建模板请求"""
    name: str = Field(..., min_length=1, max_length=100)
    description: Optional[str] = None
    fields: List[FieldRequest] = Field(..., min_items=1, description="至少包含一个字段")
    class Config:
        extra = 'forbid'


class UpdateTemplateRequest(BaseModel):
    """更新模板请求"""
    name: Optional[str] = Field(None, min_length=1, max_length=100)
    description: Optional[str] = None
    fields: Optional[List[FieldRequest]] = None
    class Config:
        extra = 'forbid'


# ==================== API 路由 ====================

@router.get("/", response_model=List[TemplateResponse])
async def get_templates(
    current_user: Secretary = Depends(get_current_user),
    db: Session = Depends(get_db_session)
):
    """
    获取当前用户创建的所有模板
    """
    templates = db.query(TemplateForm).filter(
        TemplateForm.created_by == current_user.id
    ).all()
    
    result = []
    for template in templates:
        field_count = db.query(TemplateFormField).filter(
            TemplateFormField.form_id == template.id
        ).count()
        
        result.append(TemplateResponse(
            id=template.id,
            name=template.name,
            description=template.description,
            created_at=ensure_utc(template.created_at),
            field_count=field_count
        ))
    
    return result


@router.get("/list", response_model=List[TemplateResponse])
async def get_templates_list(
    current_user: Secretary = Depends(get_current_user),
    db: Session = Depends(get_db_session)
):
    """
    获取当前用户创建的所有模板（别名路由）
    """
    return await get_templates(current_user, db)


@router.get("/{template_id}", response_model=TemplateDetailResponse)
async def get_template_detail(
    template_id: int,
    current_user: Secretary = Depends(get_current_user),
    db: Session = Depends(get_db_session)
):
    """
    获取模板详情，包含所有字段
    """
    template = db.query(TemplateForm).filter(
        TemplateForm.id == template_id,
        TemplateForm.created_by == current_user.id
    ).first()
    
    if not template:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="模板不存在或无权访问"
        )
    
    fields = db.query(TemplateFormField).filter(
        TemplateFormField.form_id == template_id
    ).order_by(TemplateFormField.ord).all()
    
    return TemplateDetailResponse(
        id=template.id,
        name=template.name,
        description=template.description,
        created_at=ensure_utc(template.created_at),
        fields=[
            FieldResponse(
                id=field.id,
                display_name=field.display_name,
                validation_rule=field.validation_rule,
                ord=field.ord
            )
            for field in fields
        ]
    )


@router.post("/create")
async def create_template(
    request: CreateTemplateRequest,
    current_user: Secretary = Depends(get_current_user),
    db: Session = Depends(get_db_session)
):
    """
    创建新模板
    """
    # 将 Pydantic 模型转换为字典列表
    fields_data = [
        {
            "display_name": field.display_name,
            "validation_rule": field.validation_rule,
            "ord": field.ord
        }
        for field in request.fields
    ]
    
    # 调用核心业务逻辑
    result = create_template_core(
        name=request.name,
        fields=fields_data,
        description=request.description,
        created_by=current_user.id,
        db=db
    )
    
    # 根据结果返回 HTTP 响应
    if result["success"]:
        return result
    else:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=result["message"]
        )


@router.put("/{template_id}")
async def update_template(
    template_id: int,
    request: UpdateTemplateRequest,
    current_user: Secretary = Depends(get_current_user),
    db: Session = Depends(get_db_session)
):
    """
    更新模板
    """
    # 将 Pydantic 模型转换为字典列表
    fields_data = None
    if request.fields is not None:
        fields_data = [
            {
                "display_name": field.display_name,
                "validation_rule": field.validation_rule,
                "ord": field.ord
            }
            for field in request.fields
        ]
    
    # 调用核心业务逻辑
    result = update_template_core(
        template_id=template_id,
        name=request.name,
        fields=fields_data,
        description=request.description,
        user_id=current_user.id,
        db=db
    )
    
    # 根据结果返回 HTTP 响应
    if result["success"]:
        return result
    else:
        # 根据错误信息决定状态码
        if "不存在" in result["message"] or "无权访问" in result["message"]:
            status_code = status.HTTP_404_NOT_FOUND
        else:
            status_code = status.HTTP_400_BAD_REQUEST
        
        raise HTTPException(
            status_code=status_code,
            detail=result["message"]
        )


@router.delete("/{template_id}")
async def delete_template(
    template_id: int,
    current_user: Secretary = Depends(get_current_user),
    db: Session = Depends(get_db_session)
):
    """
    删除模板
    """
    template = db.query(TemplateForm).filter(
        TemplateForm.id == template_id,
        TemplateForm.created_by == current_user.id
    ).first()
    
    if not template:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="模板不存在或无权访问"
        )
    
    try:
        # 删除所有字段（级联）
        db.query(TemplateFormField).filter(
            TemplateFormField.form_id == template_id
        ).delete()
        
        # 删除模板
        db.delete(template)
        db.commit()
        
        return {
            "success": True,
            "message": "模板删除成功"
        }
    
    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"删除模板失败：{str(e)}"
        )

@router.post("/parse-excel")
async def parse_excel(
    file: UploadFile = File(...),
    current_user: Secretary = Depends(get_current_user)
):
    """
    解析 Excel 文件，提取表头作为模板字段
    """
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="仅支持 .xlsx 和 .xls 格式的 Excel 文件"
        )
    
    try:
        # 读取文件内容
        contents = await file.read()
        workbook = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
        
        # 获取第一个工作表
        sheet = workbook.active
        
        # 获取文件名作为模板名称（去除扩展名）
        template_name = file.filename.rsplit('.', 1)[0]
        
        # 读取第一行作为字段名
        first_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=False))
        
        fields = []
        for idx, cell in enumerate(first_row):
            if cell.value is None or str(cell.value).strip() == '':
                continue
                
            field_name = str(cell.value).strip()
            
            # 提取 DataValidation 规则（只当 Excel 中有规则才返回），否则 None
            validation_rule = extract_column_validation(sheet, idx + 1)
            fields.append({
                'display_name': field_name,
                'validation_rule': validation_rule,
                'ord': idx
            })
        
        if not fields:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Excel 文件第一行没有有效的列名"
            )
        
        return {
            "template_name": template_name,
            "fields": fields
        }
    
    except openpyxl.utils.exceptions.InvalidFileException:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="无效的 Excel 文件格式"
        )
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"解析 Excel 失败：{str(e)}"
        )

def extract_column_validation(sheet, column_index: int) -> dict:
    """
    Extract a validation_rule dict from openpyxl sheet for a column (based on header cell/column validations)
    Returns a validation_rule JSON structure or None
    """
    rule = None

    try:
        dv_list = list(sheet.data_validations.dataValidation) if sheet.data_validations is not None else []
        # Look for a validation that targets rows in this column (we expect sqref to contain column like 'A:A' or 'A2:A100')
        for dv in dv_list:
            if not dv.sqref:
                continue
            ranges = []
            # openpyxl may store sqref as e.g. 'A1:A100' or 'A:A'
            for r in str(dv.sqref).split():
                ranges.append(r)
            for r in ranges:
                # Only check simple column ranges like A, A:A, A1:A100
                if r.startswith(' '):
                    r = r.strip()
                # Get column letter for first cell in range
                if ':' in r:
                    start_cell = r.split(':')[0]
                else:
                    start_cell = r
                # Extract column letter(s)
                col_letter = ''.join(ch for ch in start_cell if ch.isalpha())
                if not col_letter:
                    continue
                # Convert column letter to index
                from openpyxl.utils import column_index_from_string
                try:
                    dv_col_index = column_index_from_string(col_letter)
                except Exception:
                    continue
                if dv_col_index != column_index:
                    continue

                # Found a DataValidation for this column
                # Map validation types
                vtype = (dv.type or '').lower()
                if vtype == 'list':
                    # formula1 may be a quoted comma list or a range; handle simple quoted list
                    formula = dv.formula1 or ''
                    opts = []
                    if formula.startswith('"') and formula.endswith('"'):
                        inner = formula.strip('"')
                        opts = [s.strip() for s in inner.split(',') if s.strip()]
                    # Do not create SELECT type; keep TEXT with options
                    rule = rule or {}
                    rule['type'] = 'TEXT'
                    if opts:
                        rule['options'] = opts
                elif vtype in ('whole', 'decimal'):
                    rule = rule or {}
                    # 'whole' -> INTEGER, 'decimal' -> FLOAT
                    rule['type'] = 'INTEGER' if vtype == 'whole' else 'FLOAT'
                    if dv.formula1:
                        try:
                            rule['min'] = float(dv.formula1)
                        except Exception:
                            pass
                    if dv.formula2:
                        try:
                            rule['max'] = float(dv.formula2)
                        except Exception:
                            pass
                elif vtype == 'date':
                    rule = rule or {}
                    rule['type'] = 'DATE'
                    # We won't parse min/max here; leave as type
                elif vtype == 'textLength':
                    rule = rule or {}
                    rule['type'] = 'TEXT'
                    if dv.operator in ('between',):
                        try:
                            rule['min_length'] = int(dv.formula1)
                            rule['max_length'] = int(dv.formula2)
                        except Exception:
                            pass
                elif vtype == 'custom':
                    # For custom we may have formula like =ISNUMBER(SEARCH("@",A2)) etc — leave it as custom regex not parsed
                    rule = rule or {}
                    rule.setdefault('extra', {})['custom_formula'] = dv.formula1
                # We don't set required since Excel doesn't have explicit required VIA DataValidation
                # Break after first match
                return rule
    except Exception:
        pass

    # If we did not find any relevant DataValidation for this column, return None to indicate no rule
    return rule


# def infer_field_type(sheet, column_index: int) -> str:
#     """
#     根据列的数据推断字段类型
#     检查前10行数据（跳过表头）
#     """
#     # 默认类型
#     default_type = 'TEXT'
    
#     # 收集样本数据
#     samples = []
#     for row in sheet.iter_rows(min_row=2, max_row=11, min_col=column_index, max_col=column_index):
#         cell = row[0]
#         if cell.value is not None and str(cell.value).strip():
#             samples.append(cell)
    
#     if not samples:
#         return default_type
    
#     # 检查是否所有样本都是数字
#     all_numbers = all(_is_number(cell.value) for cell in samples)
#     if all_numbers:
#         # Check whether all are integers
#         all_int = all(float(cell.value).is_integer() for cell in samples)
#         return 'INTEGER' if all_int else 'FLOAT'
    
#     # 检查是否有日期格式
#     has_date = any(_is_date_cell(cell) for cell in samples)
#     if has_date:
#         return 'DATE'
    
#     # 检查是否有时间格式
#     has_time = any(_is_time_cell(cell) for cell in samples)
#     if has_time:
#         return 'TIME'
    
#     # 检查是否是布尔值
#     all_boolean = all(_is_boolean(cell.value) for cell in samples)
#     if all_boolean:
#         return 'BOOLEAN'
    
#     # 检查是否是选项（重复值较少）
#     unique_values = set(str(cell.value).strip() for cell in samples)
#     if len(unique_values) <= 5 and len(samples) >= 5:
#         # 如果唯一值较少，可能是单选或多选
#         return 'RADIO'
    
#     return default_type


# def _is_number(value) -> bool:
#     """判断是否为数字"""
#     try:
#         float(value)
#         return True
#     except (ValueError, TypeError):
#         return False


# def _is_date_cell(cell: Cell) -> bool:
#     """判断是否为日期单元格"""
#     if cell.is_date:
#         return True
    
#     # 检查格式代码
#     if cell.number_format:
#         date_formats = ['yyyy', 'yy', 'mm', 'dd', 'd/m', 'm/d']
#         format_lower = cell.number_format.lower()
#         return any(fmt in format_lower for fmt in date_formats)
    
#     return False


# def _is_time_cell(cell: Cell) -> bool:
#     """判断是否为时间单元格"""
#     if cell.number_format:
#         time_formats = ['h:mm', 'hh:mm', 'h:mm:ss']
#         format_lower = cell.number_format.lower()
#         return any(fmt in format_lower for fmt in time_formats)
    
#     return False


# def _is_boolean(value) -> bool:
#     """判断是否为布尔值"""
#     if isinstance(value, bool):
#         return True
    
#     if isinstance(value, str):
#         value_lower = value.lower().strip()
#         return value_lower in ['true', 'false', 'yes', 'no', '是', '否', '对', '错', '1', '0']
    
#     return False
